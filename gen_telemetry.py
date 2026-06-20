import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

START = datetime(2026, 5, 1, 0, 0, 0)
INTERVAL_MIN = 5
DAYS = 6
N = int(DAYS * 24 * 60 / INTERVAL_MIN)
ORBIT_PERIOD_MIN = 92  # typical LEO cubesat orbit

rows = []
orbit_num = 1
last_phase = 0
gs_lat, gs_lon = 20.27, 85.84  # Bhubaneswar ground station

battery_pct = 85.0
degrade = 0.0

for i in range(N):
    t = START + timedelta(minutes=i * INTERVAL_MIN)
    phase = (i * INTERVAL_MIN % ORBIT_PERIOD_MIN) / ORBIT_PERIOD_MIN  # 0-1
    if phase < last_phase:
        orbit_num += 1
    last_phase = phase

    # eclipse roughly 35% of orbit
    in_sun = phase > 0.35

    # temperature: cold in eclipse, hot in sun, sinusoidal + noise
    if in_sun:
        temp = 28 + 14 * np.sin((phase - 0.35) / 0.65 * np.pi) + np.random.normal(0, 1.5)
    else:
        temp = -12 - 8 * np.sin(phase / 0.35 * np.pi) + np.random.normal(0, 1.2)

    # battery: charges in sun, drains in eclipse, slow long-term degradation
    degrade += 0.0007
    if in_sun:
        battery_pct += np.random.uniform(0.6, 1.1)
    else:
        battery_pct -= np.random.uniform(0.5, 0.9)
    battery_pct = min(99.5, max(18, battery_pct)) - degrade * 0  # keep simple, no permanent loss
    battery_voltage = round(3.0 + (battery_pct / 100) * 1.3 + np.random.normal(0, 0.02), 3)

    # ground station contact: short windows, a few times a day
    gs_contact = phase < 0.06 or (0.5 < phase < 0.56)
    if gs_contact:
        signal = -55 - 35 * (phase % 0.06) / 0.06 + np.random.normal(0, 3)
    else:
        signal = None

    # simple simulated ground track
    altitude_km = round(500 + np.random.normal(0, 2.5), 2)
    lat = round(51.6 * np.sin(2 * np.pi * (phase + orbit_num * 0.13)), 4)
    lon = round(((i * 4.5) % 360) - 180, 4)

    # anomalies: rare events worth flagging in the dashboard
    anomaly = None
    if np.random.rand() < 0.004:
        spike = np.random.choice(["temp_spike", "voltage_dip", "signal_dropout"])
        if spike == "temp_spike":
            temp += np.random.uniform(15, 25)
        elif spike == "voltage_dip":
            battery_voltage -= np.random.uniform(0.3, 0.5)
        else:
            signal = (signal or -90) - np.random.uniform(15, 25)
        anomaly = spike

    rows.append({
        "Timestamp": t,
        "Orbit_Number": orbit_num,
        "Orbit_Phase_Pct": round(phase * 100, 1),
        "In_Sunlight": in_sun,
        "Temperature_C": round(temp, 2),
        "Battery_Voltage_V": battery_voltage,
        "Battery_Percent": round(battery_pct, 1),
        "Ground_Station_Contact": gs_contact,
        "Signal_Strength_dBm": round(signal, 1) if signal is not None else None,
        "Altitude_km": altitude_km,
        "Latitude": lat,
        "Longitude": lon,
        "Anomaly_Flag": anomaly is not None,
        "Anomaly_Type": anomaly if anomaly else "",
    })

df = pd.DataFrame(rows)

out_path = "/home/claude/CubeSat_Telemetry_Data.xlsx"
df.to_excel(out_path, index=False, sheet_name="Telemetry_Raw")

wb = load_workbook(out_path)
ws = wb["Telemetry_Raw"]

header_fill = PatternFill("solid", start_color="1F2A44", end_color="1F2A44")
header_font = Font(bold=True, color="FFFFFF", name="Arial")
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)

ws.freeze_panes = "A2"
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).font = Font(name="Arial", size=10)

# Data dictionary sheet
dd = wb.create_sheet("Data_Dictionary")
dd_rows = [
    ("Column", "Description"),
    ("Timestamp", "UTC time of telemetry reading, logged every 5 minutes"),
    ("Orbit_Number", "Sequential orbit count since simulation start"),
    ("Orbit_Phase_Pct", "Position within current orbit, 0-100%"),
    ("In_Sunlight", "Whether the satellite is in direct sunlight (vs eclipse)"),
    ("Temperature_C", "Onboard temperature sensor reading, Celsius"),
    ("Battery_Voltage_V", "Battery pack voltage"),
    ("Battery_Percent", "Estimated battery state of charge"),
    ("Ground_Station_Contact", "Whether satellite is within ground station communication window"),
    ("Signal_Strength_dBm", "Downlink signal strength during contact windows only (blank otherwise)"),
    ("Altitude_km", "Orbital altitude"),
    ("Latitude / Longitude", "Simulated ground track position"),
    ("Anomaly_Flag", "TRUE if an anomalous reading was detected this interval"),
    ("Anomaly_Type", "temp_spike / voltage_dip / signal_dropout"),
]
for r in dd_rows:
    dd.append(r)
dd.column_dimensions["A"].width = 26
dd.column_dimensions["B"].width = 75
for row in dd.iter_rows(min_row=2, max_row=dd.max_row):
    for cell in row:
        cell.font = Font(name="Arial", size=10)
dd["A1"].font = Font(bold=True, name="Arial")
dd["B1"].font = Font(bold=True, name="Arial")

# Quick summary sheet with live Excel formulas (also doubles as Excel-skill practice)
sm = wb.create_sheet("Quick_Summary")
n_rows = df.shape[0]
last_row = n_rows + 1  # +1 for header in Telemetry_Raw
summary_items = [
    ("Metric", "Value", "Formula Used"),
    ("Total Readings", f"=COUNTA(Telemetry_Raw!A2:A{last_row})", "COUNTA"),
    ("Average Temperature (C)", f"=AVERAGE(Telemetry_Raw!E2:E{last_row})", "AVERAGE"),
    ("Max Temperature (C)", f"=MAX(Telemetry_Raw!E2:E{last_row})", "MAX"),
    ("Min Temperature (C)", f"=MIN(Telemetry_Raw!E2:E{last_row})", "MIN"),
    ("Average Battery (%)", f"=AVERAGE(Telemetry_Raw!G2:G{last_row})", "AVERAGE"),
    ("Min Battery (%)", f"=MIN(Telemetry_Raw!G2:G{last_row})", "MIN"),
    ("Total Anomalies Detected", f"=COUNTIF(Telemetry_Raw!M2:M{last_row},TRUE)", "COUNTIF"),
    ("Ground Station Contacts", f"=COUNTIF(Telemetry_Raw!H2:H{last_row},TRUE)", "COUNTIF"),
    ("Anomaly Rate (%)", f"=ROUND(D9/D2*100,2)", "ROUND / cell math"),
]
for r in summary_items:
    sm.append(r)
for cell in sm[1]:
    cell.font = Font(bold=True, name="Arial")
    cell.fill = PatternFill("solid", start_color="1F2A44", end_color="1F2A44")
    cell.font = Font(bold=True, color="FFFFFF", name="Arial")
sm.column_dimensions["A"].width = 28
sm.column_dimensions["B"].width = 18
sm.column_dimensions["C"].width = 22
for row in sm.iter_rows(min_row=2, max_row=sm.max_row):
    for cell in row:
        cell.font = Font(name="Arial", size=10)

# fix the anomaly rate formula reference (D9 = anomalies row, D2 = total readings row)
sm["B10"] = "=ROUND(B9/B2*100,2)"

wb.move_sheet("Telemetry_Raw", offset=-(wb.sheetnames.index("Telemetry_Raw")))
wb.save(out_path)
print(f"rows={n_rows}")
